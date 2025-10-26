"""
Script để tạo 9 file Excel cho 9 tables trong Data Warehouse
Format theo mẫu: No | Field name | key | data type | constraints | default | description | ex
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Tạo folder nếu chưa có
output_folder = 'd:/script/table_designs'
os.makedirs(output_folder, exist_ok=True)

# Style cho header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_excel_table(filename, table_name, columns_data):
    """
    Tạo file Excel với format chuẩn
    
    columns_data: list of dict với keys:
        - field_name
        - key (PK, FK, hoặc '')
        - data_type
        - constraints
        - default
        - description
        - ex (example)
    """
    
    # Tạo DataFrame
    data = []
    for idx, col in enumerate(columns_data, 1):
        data.append({
            'No': idx,
            'Field name': col.get('field_name', ''),
            'key': col.get('key', ''),
            'data type': col.get('data_type', ''),
            'constraints': col.get('constraints', ''),
            'default': col.get('default', ''),
            'description': col.get('description', ''),
            'ex': col.get('ex', '')
        })
    
    df = pd.DataFrame(data)
    
    # Tạo Excel writer
    filepath = os.path.join(output_folder, filename)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Table Design', index=False)
        
        # Get workbook và worksheet
        workbook = writer.book
        worksheet = writer.sheets['Table Design']
        
        # Add table name row
        worksheet.insert_rows(1)
        worksheet['A1'] = f'Table name: {table_name}'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:H1')
        
        # Style header row (now row 2)
        for cell in worksheet[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Style data rows
        for row in worksheet.iter_rows(min_row=3, max_row=len(data)+2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # Adjust column widths
        column_widths = {
            'A': 5,   # No
            'B': 25,  # Field name
            'C': 8,   # key
            'D': 18,  # data type
            'E': 20,  # constraints
            'F': 15,  # default
            'G': 30,  # description
            'H': 25   # ex
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Set row heights
        worksheet.row_dimensions[1].height = 25  # Title row
        worksheet.row_dimensions[2].height = 30  # Header row
        
        for row in range(3, len(data)+3):
            worksheet.row_dimensions[row].height = 20
    
    print(f"✅ Created: {filename}")


# ============================================
# 1. FACT TABLE: fact_book_sales
# ============================================

fact_book_sales_columns = [
    {
        'field_name': 'fact_id',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Mã định danh fact (surrogate key)',
        'ex': '1'
    },
    {
        'field_name': 'date_key',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Khóa ngoại đến dim_date',
        'ex': '20241016'
    },
    {
        'field_name': 'product_key',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Khóa ngoại đến dim_product',
        'ex': '5001'
    },
    {
        'field_name': 'author_key',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Khóa ngoại đến dim_author',
        'ex': '3001'
    },
    {
        'field_name': 'publisher_key',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Khóa ngoại đến dim_publisher',
        'ex': '4001'
    },
    {
        'field_name': 'category_key',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Khóa ngoại đến dim_category',
        'ex': '6001'
    },
    {
        'field_name': 'supplier_key',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Khóa ngoại đến dim_supplier',
        'ex': '7001'
    },
    {
        'field_name': 'quantity_sold',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Số lượng đã bán',
        'ex': '100'
    },
    {
        'field_name': 'review_count',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Số lượng đánh giá',
        'ex': '1234'
    },
    {
        'field_name': 'revenue',
        'key': '',
        'data_type': 'DECIMAL(15,2)',
        'constraints': '',
        'default': '',
        'description': 'Doanh thu (quantity_sold * discount_price)',
        'ex': '12000000.00'
    },
    {
        'field_name': 'original_price',
        'key': '',
        'data_type': 'DECIMAL(15,2)',
        'constraints': '',
        'default': '',
        'description': 'Giá gốc',
        'ex': '150000.00'
    },
    {
        'field_name': 'discount_price',
        'key': '',
        'data_type': 'DECIMAL(15,2)',
        'constraints': '',
        'default': '',
        'description': 'Giá sau giảm',
        'ex': '120000.00'
    },
    {
        'field_name': 'discount_percent',
        'key': '',
        'data_type': 'DECIMAL(5,2)',
        'constraints': 'CHECK >= 0 AND <= 100',
        'default': '',
        'description': 'Tỷ lệ giảm giá (%)',
        'ex': '20.00'
    },
    {
        'field_name': 'rating',
        'key': '',
        'data_type': 'DECIMAL(3,2)',
        'constraints': 'CHECK >= 0 AND <= 5',
        'default': '',
        'description': 'Đánh giá trung bình (0-5 sao)',
        'ex': '4.8'
    },
    {
        'field_name': 'is_bestseller',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': 'FALSE',
        'description': 'Sách bán chạy?',
        'ex': 'TRUE'
    },
    {
        'field_name': 'is_on_sale',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': 'FALSE',
        'description': 'Đang giảm giá?',
        'ex': 'TRUE'
    },
    {
        'field_name': 'url_path',
        'key': '',
        'data_type': 'VARCHAR(500)',
        'constraints': '',
        'default': '',
        'description': 'URL sản phẩm khi crawl',
        'ex': 'https://fahasa.com/dac-nhan-tam.html'
    },
    {
        'field_name': 'crawl_timestamp',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Thời điểm crawl dữ liệu',
        'ex': '2024-10-16 14:30:00'
    },
    {
        'field_name': 'elt_batch_id',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': 'NULL',
        'description': 'Khóa ngoại đến crawl_log (optional)',
        'ex': '101'
    },
    {
        'field_name': 'data_src',
        'key': '',
        'data_type': 'VARCHAR(50)',
        'constraints': '',
        'default': 'Fahasa.com',
        'description': 'Nguồn dữ liệu',
        'ex': 'Fahasa.com'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('01_fact_book_sales.xlsx', 'fact_book_sales', fact_book_sales_columns)


# ============================================
# 2. DIMENSION: dim_date
# ============================================

dim_date_columns = [
    {
        'field_name': 'date_key',
        'key': 'PK',
        'data_type': 'INTEGER',
        'constraints': 'NOT NULL, UNIQUE',
        'default': '',
        'description': 'Khóa chính (YYYYMMDD format)',
        'ex': '20241016'
    },
    {
        'field_name': 'full_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': 'NOT NULL, UNIQUE',
        'default': '',
        'description': 'Ngày đầy đủ',
        'ex': '2024-10-16'
    },
    {
        'field_name': 'day_of_week',
        'key': '',
        'data_type': 'VARCHAR(10)',
        'constraints': '',
        'default': '',
        'description': 'Thứ trong tuần (English)',
        'ex': 'Wednesday'
    },
    {
        'field_name': 'day_of_month',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Ngày trong tháng (1-31)',
        'ex': '16'
    },
    {
        'field_name': 'day_of_year',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Ngày trong năm (1-365)',
        'ex': '290'
    },
    {
        'field_name': 'week_start_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': '',
        'default': '',
        'description': 'Ngày đầu tuần (Monday)',
        'ex': '2024-10-14'
    },
    {
        'field_name': 'week_end_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': '',
        'default': '',
        'description': 'Ngày cuối tuần (Sunday)',
        'ex': '2024-10-20'
    },
    {
        'field_name': 'month',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': 'CHECK >= 1 AND <= 12',
        'default': '',
        'description': 'Tháng (1-12)',
        'ex': '10'
    },
    {
        'field_name': 'month_name',
        'key': '',
        'data_type': 'VARCHAR(10)',
        'constraints': '',
        'default': '',
        'description': 'Tên tháng (English)',
        'ex': 'October'
    },
    {
        'field_name': 'month_year',
        'key': '',
        'data_type': 'VARCHAR(7)',
        'constraints': '',
        'default': '',
        'description': 'Tháng-Năm (YYYY-MM)',
        'ex': '2024-10'
    },
    {
        'field_name': 'week_of_year',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Tuần trong năm (1-52)',
        'ex': '42'
    },
    {
        'field_name': 'day_of_week_vn',
        'key': '',
        'data_type': 'VARCHAR(15)',
        'constraints': '',
        'default': '',
        'description': 'Thứ trong tuần (Tiếng Việt)',
        'ex': 'Thứ Tư'
    },
    {
        'field_name': 'is_last_day_of_month',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': '',
        'description': 'Ngày cuối tháng?',
        'ex': 'FALSE'
    },
    {
        'field_name': 'is_first_day_of_month',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': '',
        'description': 'Ngày đầu tháng?',
        'ex': 'FALSE'
    },
    {
        'field_name': 'quarter',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': 'CHECK >= 1 AND <= 4',
        'default': '',
        'description': 'Quý (1-4)',
        'ex': '4'
    },
    {
        'field_name': 'quarter_name',
        'key': '',
        'data_type': 'VARCHAR(10)',
        'constraints': '',
        'default': '',
        'description': 'Tên quý',
        'ex': 'Q4'
    },
    {
        'field_name': 'year',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Năm',
        'ex': '2024'
    },
    {
        'field_name': 'is_weekend',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': '',
        'description': 'Cuối tuần? (T7, CN)',
        'ex': 'FALSE'
    },
    {
        'field_name': 'is_holiday',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': '',
        'description': 'Ngày lễ?',
        'ex': 'FALSE'
    },
    {
        'field_name': 'is_workday',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': '',
        'description': 'Ngày làm việc?',
        'ex': 'TRUE'
    },
    {
        'field_name': 'holiday_name',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': 'NULL',
        'description': 'Tên ngày lễ (nếu có)',
        'ex': 'Tết Nguyên Đán'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('02_dim_date.xlsx', 'dim_date', dim_date_columns)


# ============================================
# 3. DIMENSION: dim_product
# ============================================

dim_product_columns = [
    {
        'field_name': 'product_key',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '5001'
    },
    {
        'field_name': 'product_id',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Mã sản phẩm từ Fahasa (natural key)',
        'ex': 'dac-nhan-tam-8935086855362'
    },
    {
        'field_name': 'product_name',
        'key': '',
        'data_type': 'VARCHAR(500)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Tên sản phẩm',
        'ex': 'Đắc Nhân Tâm'
    },
    {
        'field_name': 'publish_year',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Năm xuất bản',
        'ex': '2020'
    },
    {
        'field_name': 'language',
        'key': '',
        'data_type': 'VARCHAR(50)',
        'constraints': '',
        'default': 'Tiếng Việt',
        'description': 'Ngôn ngữ',
        'ex': 'Tiếng Việt'
    },
    {
        'field_name': 'page_count',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số trang',
        'ex': '320'
    },
    {
        'field_name': 'weight',
        'key': '',
        'data_type': 'DECIMAL(10,2)',
        'constraints': '',
        'default': '',
        'description': 'Trọng lượng (kg)',
        'ex': '0.35'
    },
    {
        'field_name': 'dimensions',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': '',
        'description': 'Kích thước (cm)',
        'ex': '14x20.5cm'
    },
    {
        'field_name': 'image_url',
        'key': '',
        'data_type': 'VARCHAR(500)',
        'constraints': '',
        'default': '',
        'description': 'URL ảnh bìa sách',
        'ex': 'https://cdn.fahasa.com/media/catalog/product/d/a/dac-nhan-tam.jpg'
    },
    {
        'field_name': 'product_url',
        'key': '',
        'data_type': 'VARCHAR(500)',
        'constraints': '',
        'default': '',
        'description': 'URL trang chi tiết sản phẩm',
        'ex': 'https://fahasa.com/dac-nhan-tam.html'
    },
    {
        'field_name': 'effective_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': 'NOT NULL',
        'default': 'CURRENT_DATE',
        'description': 'Ngày bắt đầu version này (SCD Type 2)',
        'ex': '2024-01-01'
    },
    {
        'field_name': 'expiration_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': 'NOT NULL',
        'default': '9999-12-31',
        'description': 'Ngày kết thúc version (9999-12-31 if current)',
        'ex': '9999-12-31'
    },
    {
        'field_name': 'is_current',
        'key': '',
        'data_type': 'BOOLEAN',
        'constraints': '',
        'default': 'TRUE',
        'description': 'Version hiện tại?',
        'ex': 'TRUE'
    },
    {
        'field_name': 'version_number',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': 'CHECK > 0',
        'default': '1',
        'description': 'Số version',
        'ex': '1'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    },
    {
        'field_name': 'updated_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': '',
        'description': 'Thời gian cập nhật',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('03_dim_product.xlsx', 'dim_product', dim_product_columns)


# ============================================
# 4. DIMENSION: dim_author
# ============================================

dim_author_columns = [
    {
        'field_name': 'author_key',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '3001'
    },
    {
        'field_name': 'author_id',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': 'UNIQUE',
        'default': '',
        'description': 'Mã tác giả từ Fahasa (natural key)',
        'ex': 'dale-carnegie'
    },
    {
        'field_name': 'author_name',
        'key': '',
        'data_type': 'VARCHAR(200)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Tên tác giả',
        'ex': 'Dale Carnegie'
    },
    {
        'field_name': 'author_name_normalized',
        'key': '',
        'data_type': 'VARCHAR(200)',
        'constraints': '',
        'default': '',
        'description': 'Tên tác giả chuẩn hóa (lowercase, no accents)',
        'ex': 'dale carnegie'
    },
    {
        'field_name': 'total_books',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Tổng số sách (aggregated)',
        'ex': '15'
    },
    {
        'field_name': 'total_sold',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Tổng số lượng bán (aggregated)',
        'ex': '50000'
    },
    {
        'field_name': 'avg_rating',
        'key': '',
        'data_type': 'DECIMAL(3,2)',
        'constraints': '',
        'default': '',
        'description': 'Đánh giá trung bình',
        'ex': '4.5'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    },
    {
        'field_name': 'updated_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': '',
        'description': 'Thời gian cập nhật',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('04_dim_author.xlsx', 'dim_author', dim_author_columns)


# ============================================
# 5. DIMENSION: dim_publisher
# ============================================

dim_publisher_columns = [
    {
        'field_name': 'publisher_key',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '4001'
    },
    {
        'field_name': 'publisher_id',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': 'NOT NULL, UNIQUE',
        'default': '',
        'description': 'Mã NXB từ Fahasa (natural key)',
        'ex': 'nxb-tong-hop'
    },
    {
        'field_name': 'publisher_name',
        'key': '',
        'data_type': 'VARCHAR(200)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Tên nhà xuất bản',
        'ex': 'NXB Tổng Hợp TPHCM'
    },
    {
        'field_name': 'total_books',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Tổng số sách (aggregated)',
        'ex': '500'
    },
    {
        'field_name': 'avg_rating',
        'key': '',
        'data_type': 'DECIMAL(3,2)',
        'constraints': '',
        'default': '',
        'description': 'Đánh giá trung bình',
        'ex': '4.3'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    },
    {
        'field_name': 'updated_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': '',
        'description': 'Thời gian cập nhật',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('05_dim_publisher.xlsx', 'dim_publisher', dim_publisher_columns)


# ============================================
# 6. DIMENSION: dim_supplier
# ============================================

dim_supplier_columns = [
    {
        'field_name': 'supplier_key',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '7001'
    },
    {
        'field_name': 'supplier_id',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': 'NOT NULL, UNIQUE',
        'default': '',
        'description': 'Mã nhà cung cấp từ Fahasa (natural key)',
        'ex': 'alphabooks'
    },
    {
        'field_name': 'supplier_name',
        'key': '',
        'data_type': 'VARCHAR(200)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Tên nhà cung cấp',
        'ex': 'Alphabooks'
    },
    {
        'field_name': 'total_products',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Tổng số sản phẩm (aggregated)',
        'ex': '1500'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('06_dim_supplier.xlsx', 'dim_supplier', dim_supplier_columns)


# ============================================
# 7. DIMENSION: dim_category
# ============================================

dim_category_columns = [
    {
        'field_name': 'category_key',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '6001'
    },
    {
        'field_name': 'category_id',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': 'NOT NULL, UNIQUE',
        'default': '',
        'description': 'Mã category từ Fahasa (natural key)',
        'ex': 'sach-van-hoc-tieu-thuyet'
    },
    {
        'field_name': 'level_1',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': '',
        'description': 'Cấp 1 (Top category)',
        'ex': 'Sách'
    },
    {
        'field_name': 'level_2',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': '',
        'description': 'Cấp 2 (Middle category)',
        'ex': 'Văn học'
    },
    {
        'field_name': 'level_3',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': '',
        'description': 'Cấp 3 (Leaf category)',
        'ex': 'Tiểu thuyết'
    },
    {
        'field_name': 'full_path',
        'key': '',
        'data_type': 'VARCHAR(500)',
        'constraints': '',
        'default': '',
        'description': 'Đường dẫn đầy đủ (denormalized)',
        'ex': 'Sách > Văn học > Tiểu thuyết'
    },
    {
        'field_name': 'parent_category',
        'key': 'FK',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': 'NULL',
        'description': 'Khóa ngoại đến category cha (self-join)',
        'ex': '1100'
    },
    {
        'field_name': 'category_level',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': 'CHECK >= 1 AND <= 5',
        'default': '',
        'description': 'Cấp độ trong hierarchy (1-5)',
        'ex': '3'
    },
    {
        'field_name': 'total_products',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '0',
        'description': 'Tổng số sản phẩm (aggregated)',
        'ex': '250'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 14:30:00'
    }
]

create_excel_table('07_dim_category.xlsx', 'dim_category', dim_category_columns)


# ============================================
# 8. TRACKING: crawl_log
# ============================================

crawl_log_columns = [
    {
        'field_name': 'log_id',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '101'
    },
    {
        'field_name': 'crawl_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Ngày crawl',
        'ex': '2024-10-16'
    },
    {
        'field_name': 'start_time',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Thời gian bắt đầu crawl',
        'ex': '2024-10-16 14:00:00'
    },
    {
        'field_name': 'end_time',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Thời gian kết thúc crawl',
        'ex': '2024-10-16 15:30:00'
    },
    {
        'field_name': 'duration_seconds',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Thời gian chạy (giây)',
        'ex': '5400'
    },
    {
        'field_name': 'category_crawled',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': '',
        'description': 'Category được crawl',
        'ex': 'Sách > Văn học'
    },
    {
        'field_name': 'page_range',
        'key': '',
        'data_type': 'VARCHAR(50)',
        'constraints': '',
        'default': '',
        'description': 'Phạm vi trang crawl',
        'ex': '1-10'
    },
    {
        'field_name': 'total_found',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Tổng số URLs/products tìm thấy',
        'ex': '200'
    },
    {
        'field_name': 'total_inserted',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số record INSERT mới',
        'ex': '150'
    },
    {
        'field_name': 'total_updated',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số record UPDATE',
        'ex': '45'
    },
    {
        'field_name': 'total_failed',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số record lỗi',
        'ex': '5'
    },
    {
        'field_name': 'records_per_second',
        'key': '',
        'data_type': 'DECIMAL(10,2)',
        'constraints': '',
        'default': '',
        'description': 'Tốc độ xử lý (records/second)',
        'ex': '3.70'
    },
    {
        'field_name': 'avg_processing_time',
        'key': '',
        'data_type': 'DECIMAL(10,2)',
        'constraints': '',
        'default': '',
        'description': 'Thời gian xử lý trung bình (giây/record)',
        'ex': '0.27'
    },
    {
        'field_name': 'status',
        'key': '',
        'data_type': 'VARCHAR(20)',
        'constraints': '',
        'default': '',
        'description': 'Trạng thái (SUCCESS, FAILED, PARTIAL)',
        'ex': 'SUCCESS'
    },
    {
        'field_name': 'error_message',
        'key': '',
        'data_type': 'TEXT',
        'constraints': '',
        'default': '',
        'description': 'Chi tiết lỗi (nếu có)',
        'ex': 'Connection timeout at page 5'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 15:30:00'
    }
]

create_excel_table('08_crawl_log.xlsx', 'crawl_log', crawl_log_columns)


# ============================================
# 9. TRACKING: data_quality_log
# ============================================

data_quality_log_columns = [
    {
        'field_name': 'quality_id',
        'key': 'PK',
        'data_type': 'SERIAL',
        'constraints': 'NOT NULL, UNIQUE',
        'default': 'AUTO INCREMENT',
        'description': 'Khóa chính (surrogate key)',
        'ex': '1'
    },
    {
        'field_name': 'check_date',
        'key': '',
        'data_type': 'DATE',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Ngày kiểm tra',
        'ex': '2024-10-16'
    },
    {
        'field_name': 'check_time',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Thời gian kiểm tra',
        'ex': '2024-10-16 16:00:00'
    },
    {
        'field_name': 'table_name',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Tên bảng được kiểm tra (string, not FK)',
        'ex': 'fact_book_sales'
    },
    {
        'field_name': 'rule_name',
        'key': '',
        'data_type': 'VARCHAR(200)',
        'constraints': 'NOT NULL',
        'default': '',
        'description': 'Tên rule kiểm tra',
        'ex': 'price_not_null'
    },
    {
        'field_name': 'rule_description',
        'key': '',
        'data_type': 'TEXT',
        'constraints': '',
        'default': '',
        'description': 'Mô tả rule',
        'ex': 'Check if original_price is not NULL'
    },
    {
        'field_name': 'severity',
        'key': '',
        'data_type': 'VARCHAR(20)',
        'constraints': '',
        'default': '',
        'description': 'Mức độ nghiêm trọng (CRITICAL, WARNING, INFO)',
        'ex': 'WARNING'
    },
    {
        'field_name': 'records_checked',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số record được kiểm tra',
        'ex': '1000'
    },
    {
        'field_name': 'records_passed',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số record pass',
        'ex': '950'
    },
    {
        'field_name': 'records_failed',
        'key': '',
        'data_type': 'INTEGER',
        'constraints': '',
        'default': '',
        'description': 'Số record fail',
        'ex': '50'
    },
    {
        'field_name': 'failure_rate',
        'key': '',
        'data_type': 'DECIMAL(5,2)',
        'constraints': '',
        'default': '',
        'description': 'Tỷ lệ lỗi (%)',
        'ex': '5.00'
    },
    {
        'field_name': 'sample_failed_records',
        'key': '',
        'data_type': 'TEXT',
        'constraints': '',
        'default': '',
        'description': 'Mẫu các record lỗi (JSON)',
        'ex': '[{"fact_id": 1001, "price": null}, ...]'
    },
    {
        'field_name': 'action_taken',
        'key': '',
        'data_type': 'VARCHAR(100)',
        'constraints': '',
        'default': '',
        'description': 'Hành động đã thực hiện',
        'ex': 'auto_fix'
    },
    {
        'field_name': 'created_at',
        'key': '',
        'data_type': 'TIMESTAMP',
        'constraints': '',
        'default': 'CURRENT_TIMESTAMP',
        'description': 'Thời gian tạo record',
        'ex': '2024-10-16 16:00:00'
    }
]

create_excel_table('09_data_quality_log.xlsx', 'data_quality_log', data_quality_log_columns)


# ============================================
# SUMMARY
# ============================================

print("\n" + "="*60)
print("✅ HOÀN TẤT! Đã tạo 9 file Excel:")
print("="*60)
print("1. 01_fact_book_sales.xlsx")
print("2. 02_dim_date.xlsx")
print("3. 03_dim_product.xlsx")
print("4. 04_dim_author.xlsx")
print("5. 05_dim_publisher.xlsx")
print("6. 06_dim_supplier.xlsx")
print("7. 07_dim_category.xlsx")
print("8. 08_crawl_log.xlsx")
print("9. 09_data_quality_log.xlsx")
print("="*60)
print(f"📁 Location: {output_folder}")
print("="*60)
